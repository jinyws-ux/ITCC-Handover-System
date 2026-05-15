from datetime import date, datetime
from io import BytesIO
from flask import jsonify, request
from auth import require_admin
from extensions import db
from models import Group, HypercareCheck, Schedule, Shift
from serializers import format_dt


def serialize_schedule(schedule: Schedule) -> dict:
    return {'id': schedule.id, 'date': format_dt(schedule.work_date), 'shift': schedule.shift.code if schedule.shift else '-', 'shiftId': schedule.shift_id, 'group': schedule.group.name if schedule.group else '-', 'groupId': schedule.group_id, 'startTime': schedule.start_time, 'endTime': schedule.end_time, 'members': schedule.members_text, 'remark': schedule.remark}


def serialize_hypercare(check: HypercareCheck) -> dict:
    task = check.task
    return {'id': check.id, 'date': check.check_time.date().isoformat(), 'time': check.check_time.strftime('%H:%M'), 'taskId': task.id, 'taskNo': task.task_no, 'title': task.title, 'checkItem': check.check_item, 'expectedResult': check.expected_result, 'status': check.check_status, 'system': task.system.name if task.system else '-', 'factory': task.factory.name if task.factory else '-'}


def register_calendar_routes(app):
    @app.get('/api/calendar-rich')
    def calendar_rich():
        month = request.args.get('month') or date.today().strftime('%Y-%m')
        start = datetime.strptime(month + '-01', '%Y-%m-%d').date(); end = date(start.year + (1 if start.month == 12 else 0), 1 if start.month == 12 else start.month + 1, 1)
        schedule_items = [serialize_schedule(x) for x in Schedule.query.filter(Schedule.work_date >= start, Schedule.work_date < end).order_by(Schedule.work_date.asc(), Schedule.shift_id.asc()).all()]
        hypercare_items = [serialize_hypercare(x) for x in HypercareCheck.query.filter(HypercareCheck.check_time >= datetime.combine(start, datetime.min.time()), HypercareCheck.check_time < datetime.combine(end, datetime.min.time())).order_by(HypercareCheck.check_time.asc()).all()]
        by_date = {}
        for item in schedule_items: by_date.setdefault(item['date'], {'schedules': [], 'hypercare': []})['schedules'].append(item)
        for item in hypercare_items: by_date.setdefault(item['date'], {'schedules': [], 'hypercare': []})['hypercare'].append(item)
        return jsonify({'items': schedule_items, 'hypercare': hypercare_items, 'byDate': by_date})

    @app.post('/api/admin/schedules')
    def admin_create_schedule():
        user, error = require_admin()
        if error: return error
        payload = request.get_json(silent=True) or {}
        work_date = datetime.strptime(str(payload.get('workDate','')).strip(), '%Y-%m-%d').date()
        shift = Shift.query.get(payload.get('shiftId'))
        schedule = Schedule(work_date=work_date, shift_id=payload.get('shiftId'), group_id=payload.get('groupId'), start_time=str(payload.get('startTime') or shift.default_start_time), end_time=str(payload.get('endTime') or shift.default_end_time), members_text=str(payload.get('members','')).strip(), remark=str(payload.get('remark','')).strip(), created_by=user.id if user else None)
        db.session.add(schedule); db.session.commit(); return jsonify(serialize_schedule(schedule)), 201

    @app.patch('/api/admin/schedules/<int:schedule_id>')
    def admin_patch_schedule(schedule_id: int):
        _, error = require_admin()
        if error: return error
        s = Schedule.query.get_or_404(schedule_id)
        p = request.get_json(silent=True) or {}
        s.members_text = str(p.get('members', s.members_text))
        s.start_time = str(p.get('startTime', s.start_time))
        s.end_time = str(p.get('endTime', s.end_time))
        s.remark = str(p.get('remark', s.remark))
        db.session.commit(); return jsonify(serialize_schedule(s))

    @app.delete('/api/admin/schedules/<int:schedule_id>')
    def admin_delete_schedule(schedule_id: int):
        _, error = require_admin()
        if error: return error
        s = Schedule.query.get_or_404(schedule_id)
        db.session.delete(s); db.session.commit(); return jsonify({'status': 'ok'})

    @app.post('/api/admin/schedules/import-preview')
    def import_preview():
        _, error = require_admin()
        if error: return error
        month = request.form.get('month', '')
        file = request.files.get('file')
        if not month or not file:
            return jsonify({'message': 'month and file are required'}), 400
        try:
            from openpyxl import load_workbook
        except Exception:
            return jsonify({'message': 'openpyxl is required for Excel import. Please install dependencies.'}), 500
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        ws = wb.active
        shifts = {x.code: x for x in Shift.query.filter(Shift.code.in_(['D1', 'D2', 'E'])).all()}
        warnings, schedules_map = [], {}
        headers = [ws.cell(row=1, column=c).value for c in range(2, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(row=r, column=1).value or '').strip()
            if not name: continue
            for idx, hv in enumerate(headers, start=2):
                v = str(ws.cell(row=r, column=idx).value or '').strip().upper()
                if not v: continue
                if v in {'OFF', '休', 'AL', '请假'}: 
                    warnings.append({'cell': f'{ws.cell(row=r, column=idx).coordinate}', 'value': v, 'message': 'Ignored as off/leave'})
                    continue
                if v not in shifts:
                    warnings.append({'cell': f'{ws.cell(row=r, column=idx).coordinate}', 'value': v, 'message': 'Ignored unknown shift code'})
                    continue
                if isinstance(hv, datetime): d = hv.date()
                else:
                    hs = str(hv).strip()
                    if hs.isdigit(): d = datetime.strptime(f"{month}-{int(hs):02d}", '%Y-%m-%d').date()
                    else: d = datetime.strptime(hs.replace('/', '-'), '%Y-%m-%d').date()
                key = (d.isoformat(), v)
                schedules_map.setdefault(key, []).append(name)
        out = []
        for (wd, sc), members in schedules_map.items():
            sh = shifts[sc]
            out.append({'workDate': wd, 'shiftCode': sc, 'members': members, 'startTime': sh.default_start_time, 'endTime': sh.default_end_time, 'remark': 'Imported from Excel'})
        return jsonify({'month': month, 'warnings': warnings, 'schedules': out, 'summary': {'days': len(set(x['workDate'] for x in out)), 'schedules': len(out), 'warnings': len(warnings)}})

    @app.post('/api/admin/schedules/import-confirm')
    def import_confirm():
        _, error = require_admin()
        if error: return error
        p = request.get_json(silent=True) or {}
        month = p.get('month')
        overwrite = bool(p.get('overwrite', True))
        items = p.get('items') or []
        start = datetime.strptime(month + '-01', '%Y-%m-%d').date(); end = date(start.year + (1 if start.month == 12 else 0), 1 if start.month == 12 else start.month + 1, 1)
        if overwrite:
            Schedule.query.filter(Schedule.work_date >= start, Schedule.work_date < end).delete()
        shift_map = {x.code: x.id for x in Shift.query.all()}
        default_group = Group.query.first()
        for it in items:
            wd = datetime.strptime(it['workDate'], '%Y-%m-%d').date(); sid = shift_map.get(it['shiftCode'])
            if not sid: continue
            exists = Schedule.query.filter_by(work_date=wd, shift_id=sid).first()
            if exists and not overwrite: continue
            if exists: db.session.delete(exists)
            db.session.add(Schedule(work_date=wd, shift_id=sid, group_id=default_group.id if default_group else 1, start_time=it.get('startTime', '06:00'), end_time=it.get('endTime', '16:00'), members_text=', '.join(it.get('members', [])), remark=it.get('remark', 'Imported from Excel')))
        db.session.commit(); return jsonify({'status': 'ok'})
